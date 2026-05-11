"""
JioTag review scraper.

Pulls new reviews from Google Play, Apple App Store, and YouTube, dedupes
against a Google Sheet, runs each new review through Gemini for sentiment +
topic tagging, appends results to the sheet, and writes a downloadable .xlsx
snapshot.

Designed to run on GitHub Actions on a weekly schedule.
"""

import hashlib
import json
import os
import sys
import time
from datetime import datetime, timezone

import gspread
import requests
from google.oauth2.service_account import Credentials

# --- Configuration -----------------------------------------------------------

# Apps to scrape. Update these IDs if Jio renames or relaunches the app.
PLAY_STORE_APP_ID = "com.jio.consumer.jiothings"   # JioThings on Play Store
APP_STORE_APP_ID = "1549371816"                     # JioThings on App Store (numeric ID)
APP_STORE_COUNTRY = "in"

# YouTube search queries — finds review videos, then pulls their comments.
YOUTUBE_QUERIES = ["JioTag review", "JioTag Air review", "JioTag Go review"]
YOUTUBE_MAX_VIDEOS_PER_QUERY = 5
YOUTUBE_MAX_COMMENTS_PER_VIDEO = 100

# Sheet name inside the Google Sheet.
SHEET_TAB = "reviews"

# Columns in the order they appear in the sheet.
COLUMNS = [
    "review_id", "source", "product_variant", "product_url",
    "review_date", "scrape_date", "rating", "review_title", "review_text",
    "reviewer_name", "verified_purchase", "helpful_votes",
    "app_version", "device_info", "language", "media_attached",
    "sentiment", "sentiment_confidence", "topics", "competitor_mention",
    "response_from_seller",
]

# Topics the LLM will tag against. Constrained list = consistent analytics later.
TOPIC_TAGS = [
    "battery", "accuracy", "range", "app_connectivity", "find_my_device",
    "price", "build_quality", "customer_service", "setup", "community_find",
    "compatibility", "other",
]

# --- Helpers -----------------------------------------------------------------

def now_iso():
    return datetime.now(timezone.utc).isoformat(timespec="seconds")

def make_id(source: str, native_id: str) -> str:
    """Stable unique ID per review across runs, used for dedup."""
    h = hashlib.sha1(f"{source}:{native_id}".encode()).hexdigest()[:16]
    return f"{source}_{h}"

def log(msg: str):
    print(f"[{now_iso()}] {msg}", flush=True)

# --- Scrapers ----------------------------------------------------------------

def scrape_play_store():
    """Pull recent reviews of JioThings from Google Play."""
    from google_play_scraper import Sort, reviews
    log("Scraping Play Store...")
    out = []
    try:
        result, _ = reviews(
            PLAY_STORE_APP_ID,
            lang="en",
            country="in",
            sort=Sort.NEWEST,
            count=200,
        )
        for r in result:
            out.append({
                "review_id": make_id("play_store", r["reviewId"]),
                "source": "play_store",
                "product_variant": "JioThings app",
                "product_url": f"https://play.google.com/store/apps/details?id={PLAY_STORE_APP_ID}",
                "review_date": r["at"].isoformat() if r.get("at") else "",
                "scrape_date": now_iso(),
                "rating": r.get("score", ""),
                "review_title": "",
                "review_text": r.get("content", "") or "",
                "reviewer_name": r.get("userName", "") or "",
                "verified_purchase": "",
                "helpful_votes": r.get("thumbsUpCount", 0),
                "app_version": r.get("reviewCreatedVersion", "") or "",
                "device_info": "",
                "language": "en",
                "media_attached": "",
                "response_from_seller": "yes" if r.get("replyContent") else "no",
            })
        log(f"  Play Store: {len(out)} reviews fetched")
    except Exception as e:
        log(f"  Play Store error: {e}")
    return out

def scrape_app_store():
    """Pull recent reviews from Apple App Store via the public RSS feed."""
    log("Scraping App Store...")
    out = []
    try:
        # Apple's public review RSS — no API key needed, max ~500 reviews across 10 pages.
        for page in range(1, 6):
            url = (f"https://itunes.apple.com/{APP_STORE_COUNTRY}/rss/customerreviews/"
                   f"page={page}/id={APP_STORE_APP_ID}/sortby=mostrecent/json")
            resp = requests.get(url, timeout=20)
            if resp.status_code != 200:
                break
            data = resp.json()
            entries = data.get("feed", {}).get("entry", [])
            # First entry is app metadata, skip it on page 1.
            if page == 1 and entries and "im:name" in entries[0]:
                entries = entries[1:]
            if not entries:
                break
            for e in entries:
                rid = e.get("id", {}).get("label", "")
                if not rid:
                    continue
                out.append({
                    "review_id": make_id("app_store", rid),
                    "source": "app_store",
                    "product_variant": "JioThings app",
                    "product_url": f"https://apps.apple.com/{APP_STORE_COUNTRY}/app/id{APP_STORE_APP_ID}",
                    "review_date": e.get("updated", {}).get("label", ""),
                    "scrape_date": now_iso(),
                    "rating": e.get("im:rating", {}).get("label", ""),
                    "review_title": e.get("title", {}).get("label", ""),
                    "review_text": e.get("content", {}).get("label", ""),
                    "reviewer_name": e.get("author", {}).get("name", {}).get("label", ""),
                    "verified_purchase": "",
                    "helpful_votes": e.get("im:voteSum", {}).get("label", 0),
                    "app_version": e.get("im:version", {}).get("label", ""),
                    "device_info": "iOS",
                    "language": "en",
                    "media_attached": "",
                    "response_from_seller": "",
                })
            time.sleep(0.5)
        log(f"  App Store: {len(out)} reviews fetched")
    except Exception as e:
        log(f"  App Store error: {e}")
    return out

def scrape_youtube():
    """Find JioTag review videos and pull their comments."""
    from youtube_comment_downloader import YoutubeCommentDownloader
    log("Scraping YouTube...")
    out = []
    try:
        # We use YouTube's search via the comment-downloader's helper if available,
        # otherwise just use a static list of search-result URLs from a quick web search.
        # For reliability, we use yt-dlp's flat-search to get video IDs without API key.
        import yt_dlp
        downloader = YoutubeCommentDownloader()
        ydl_opts = {"quiet": True, "extract_flat": True, "skip_download": True}
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            for query in YOUTUBE_QUERIES:
                search_url = f"ytsearch{YOUTUBE_MAX_VIDEOS_PER_QUERY}:{query}"
                info = ydl.extract_info(search_url, download=False)
                for entry in info.get("entries", []):
                    video_id = entry.get("id")
                    video_title = entry.get("title", "")
                    if not video_id:
                        continue
                    video_url = f"https://www.youtube.com/watch?v={video_id}"
                    count = 0
                    for c in downloader.get_comments_from_url(video_url, sort_by=0):
                        if count >= YOUTUBE_MAX_COMMENTS_PER_VIDEO:
                            break
                        count += 1
                        out.append({
                            "review_id": make_id("youtube", f"{video_id}:{c.get('cid', '')}"),
                            "source": "youtube",
                            "product_variant": "JioTag (general)",
                            "product_url": video_url,
                            "review_date": c.get("time", ""),
                            "scrape_date": now_iso(),
                            "rating": "",
                            "review_title": f"Comment on: {video_title}",
                            "review_text": c.get("text", "") or "",
                            "reviewer_name": c.get("author", "") or "",
                            "verified_purchase": "",
                            "helpful_votes": c.get("votes", 0),
                            "app_version": "",
                            "device_info": "",
                            "language": "en",
                            "media_attached": "no",
                            "response_from_seller": "",
                        })
                    time.sleep(0.5)
        log(f"  YouTube: {len(out)} comments fetched")
    except Exception as e:
        log(f"  YouTube error: {e}")
    return out

# --- Sentiment & topic tagging via Gemini ------------------------------------

def tag_with_llm(review_text: str) -> dict:
    """Send a review to Gemini and get back sentiment + topic tags.
    Retries on 429 rate-limit errors with exponential backoff."""
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key or not review_text.strip():
        return {"sentiment": "", "sentiment_confidence": "",
                "topics": "", "competitor_mention": ""}

    prompt = f"""Classify this product review of JioTag (a Bluetooth item tracker similar to Apple AirTag, made by Reliance Jio in India).

Review: \"\"\"{review_text[:2000]}\"\"\"

Respond with ONLY a JSON object, no prose, no markdown:
{{
  "sentiment": "positive" | "neutral" | "negative",
  "sentiment_confidence": 0.0-1.0,
  "topics": [list of applicable tags from: {TOPIC_TAGS}],
  "competitor_mention": true | false (true if AirTag, Tile, SmartTag, or other competitor is named)
}}"""

    model = "gemini-2.5-flash-lite"
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"
    body = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {
            "temperature": 0.1,
            "maxOutputTokens": 300,
            "responseMimeType": "application/json",
        },
    }
    headers = {"x-goog-api-key": api_key, "Content-Type": "application/json"}

    # Retry up to 4 times on 429 / 503; exponential backoff.
    backoff = 8
    for attempt in range(4):
        try:
            resp = requests.post(url, headers=headers, json=body, timeout=30)
            if resp.status_code in (429, 503):
                log(f"  Gemini {resp.status_code} on attempt {attempt + 1}, backing off {backoff}s")
                time.sleep(backoff)
                backoff *= 2
                continue
            resp.raise_for_status()
            data = resp.json()
            text = data["candidates"][0]["content"]["parts"][0]["text"].strip()
            if text.startswith("```"):
                text = text.split("```")[1]
                if text.startswith("json"):
                    text = text[4:]
            parsed = json.loads(text.strip())
            return {
                "sentiment": parsed.get("sentiment", ""),
                "sentiment_confidence": parsed.get("sentiment_confidence", ""),
                "topics": ",".join(parsed.get("topics", [])),
                "competitor_mention": "yes" if parsed.get("competitor_mention") else "no",
            }
        except Exception as e:
            log(f"  Gemini tagging error (attempt {attempt + 1}): {e}")
            if attempt < 3:
                time.sleep(backoff)
                backoff *= 2
            continue

    return {"sentiment": "", "sentiment_confidence": "",
            "topics": "", "competitor_mention": ""}

# --- Google Sheets I/O -------------------------------------------------------

def get_sheet():
    creds_json = os.environ["GOOGLE_CREDS_JSON"]
    sheet_id = os.environ["GOOGLE_SHEET_ID"]
    creds = Credentials.from_service_account_info(
        json.loads(creds_json),
        scopes=["https://www.googleapis.com/auth/spreadsheets"],
    )
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(sheet_id)
    try:
        ws = sh.worksheet(SHEET_TAB)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=SHEET_TAB, rows=1, cols=len(COLUMNS))
        ws.append_row(COLUMNS)
    # Ensure header row is correct.
    header = ws.row_values(1)
    if header != COLUMNS:
        ws.update("A1", [COLUMNS])
    return ws

def existing_ids(ws):
    """Return the set of review_ids already in the sheet (column A)."""
    col = ws.col_values(1)
    return set(col[1:])  # skip header

def append_reviews(ws, rows):
    if not rows:
        return
    values = [[r.get(c, "") for c in COLUMNS] for r in rows]
    ws.append_rows(values, value_input_option="RAW")

# --- xlsx export -------------------------------------------------------------

def write_xlsx(new_rows, all_rows_from_sheet, path: str):
    """Write a 3-sheet xlsx: Summary, New This Run, All Reviews."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = "Summary"

    n = len(new_rows)
    by_source = {}
    by_sentiment = {"positive": 0, "neutral": 0, "negative": 0, "": 0}
    topic_counts = {}
    negative_examples = []
    for r in new_rows:
        by_source[r["source"]] = by_source.get(r["source"], 0) + 1
        s = r.get("sentiment", "") or ""
        by_sentiment[s] = by_sentiment.get(s, 0) + 1
        for t in (r.get("topics", "") or "").split(","):
            t = t.strip()
            if t:
                topic_counts[t] = topic_counts.get(t, 0) + 1
        if r.get("sentiment") == "negative" and len(negative_examples) < 10:
            snippet = (r.get("review_text") or "").strip().replace("\n", " ")
            if len(snippet) > 250:
                snippet = snippet[:250] + "..."
            negative_examples.append({
                "source": r["source"],
                "topics": r.get("topics", ""),
                "text": snippet,
            })

    bold = Font(bold=True, size=12)
    title_font = Font(bold=True, size=16)
    header_fill = PatternFill("solid", fgColor="DDDDDD")

    ws["A1"] = "JioTag Review Scrape"
    ws["A1"].font = title_font
    ws["A2"] = f"Run date: {now_iso()}"
    ws["A3"] = f"New reviews this run: {n}"
    ws["A3"].font = bold

    row = 5
    ws.cell(row=row, column=1, value="By source").font = bold
    row += 1
    for k, v in by_source.items():
        ws.cell(row=row, column=1, value=k)
        ws.cell(row=row, column=2, value=v)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Sentiment").font = bold
    row += 1
    for label, key in [("Positive", "positive"), ("Neutral", "neutral"), ("Negative", "negative")]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=by_sentiment.get(key, 0))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Top topics").font = bold
    row += 1
    for t, c in sorted(topic_counts.items(), key=lambda x: -x[1])[:10]:
        ws.cell(row=row, column=1, value=t)
        ws.cell(row=row, column=2, value=c)
        row += 1

    if negative_examples:
        row += 1
        ws.cell(row=row, column=1, value="Negative review snippets").font = bold
        row += 1
        ws.cell(row=row, column=1, value="Source").font = bold
        ws.cell(row=row, column=2, value="Topics").font = bold
        ws.cell(row=row, column=3, value="Review").font = bold
        row += 1
        for ex in negative_examples:
            ws.cell(row=row, column=1, value=ex["source"])
            ws.cell(row=row, column=2, value=ex["topics"])
            ws.cell(row=row, column=3, value=ex["text"]).alignment = Alignment(wrap_text=True)
            row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 80

    # --- Sheet 2: New This Run ---
    ws2 = wb.create_sheet("New This Run")
    ws2.append(COLUMNS)
    for cell in ws2[1]:
        cell.font = bold
        cell.fill = header_fill
    for r in new_rows:
        ws2.append([r.get(c, "") for c in COLUMNS])
    _autosize(ws2)

    # --- Sheet 3: All Reviews (full history) ---
    ws3 = wb.create_sheet("All Reviews")
    ws3.append(COLUMNS)
    for cell in ws3[1]:
        cell.font = bold
        cell.fill = header_fill
    for r in all_rows_from_sheet:
        ws3.append([r.get(c, "") for c in COLUMNS])
    _autosize(ws3)

    wb.save(path)
    log(f"xlsx written: {path}")

def _autosize(ws):
    """Set sensible column widths based on header names."""
    from openpyxl.utils import get_column_letter
    widths = {
        "review_id": 18, "source": 12, "product_variant": 18,
        "product_url": 30, "review_date": 22, "scrape_date": 22,
        "rating": 8, "review_title": 30, "review_text": 60,
        "reviewer_name": 18, "verified_purchase": 10, "helpful_votes": 10,
        "app_version": 12, "device_info": 12, "language": 10,
        "media_attached": 10, "sentiment": 12, "sentiment_confidence": 10,
        "topics": 30, "competitor_mention": 10, "response_from_seller": 12,
    }
    for i, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col_name, 15)

def fetch_all_rows_from_sheet(ws):
    """Return all rows in the sheet as list of dicts (for inclusion in the xlsx)."""
    records = ws.get_all_records()
    return records

# --- Main --------------------------------------------------------------------

def main():
    log("=== JioTag review scrape starting ===")
    ws = get_sheet()
    seen = existing_ids(ws)
    log(f"Sheet has {len(seen)} existing reviews")

    all_scraped = []
    all_scraped += scrape_play_store()
    all_scraped += scrape_app_store()
    all_scraped += scrape_youtube()

    new_rows = [r for r in all_scraped if r["review_id"] not in seen]
    log(f"Total scraped: {len(all_scraped)} | New: {len(new_rows)}")

    # Tag with Gemini. Skip if no API key (graceful degradation).
    if os.environ.get("GEMINI_API_KEY"):
        log("Tagging new reviews with Gemini...")
        for i, r in enumerate(new_rows, 1):
            tags = tag_with_llm(r["review_text"])
            r.update(tags)
            if i % 10 == 0:
                log(f"  Tagged {i}/{len(new_rows)}")
            time.sleep(5)  # 12 req/min, well under Gemini Flash-Lite's 15 RPM free limit
    else:
        log("Skipping LLM tagging (no GEMINI_API_KEY)")

    append_reviews(ws, new_rows)
    log(f"Appended {len(new_rows)} rows to sheet")

    # Write a complete xlsx snapshot — uploaded as a GitHub Actions artifact.
    all_rows = fetch_all_rows_from_sheet(ws)
    date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    xlsx_path = f"jiotag_reviews_{date_str}.xlsx"
    write_xlsx(new_rows, all_rows, xlsx_path)
    log("=== Done ===")

if __name__ == "__main__":
    main()
